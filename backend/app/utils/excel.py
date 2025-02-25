from typing import BinaryIO, Dict, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

def read_workbook(file: BinaryIO) -> pd.DataFrame:
    """读取Excel文件"""
    try:
        df = pd.read_excel(file, engine='openpyxl')
        # 清理列名（去除空格和特殊字符）
        df.columns = df.columns.str.strip().str.replace(r'[^\w\s]', '')
        return df
    except Exception as e:
        raise ValueError(f"读取Excel文件失败: {str(e)}")

def create_workbook(data: List[Dict], template_type: str = "packing_list") -> bytes:
    """创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    
    # 设置样式
    header_font = Font(name='微软雅黑', bold=True, size=11)
    header_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if template_type == "packing_list":
        # 装箱单模板
        headers = [
            ('店铺', 15),
            ('类型', 10),
            ('SKU', 15),
            ('中文名称', 30),
            ('数量', 10),
            ('箱号', 15),
            ('装箱数量', 10),
            ('规格', 20),
            ('重量(kg)', 12),
            ('体积(m³)', 12),
            ('备注', 20)
        ]
        
        # 设置表头
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 填充数据
        row = 2
        for item in data:
            for box in item.get('box_quantities', []):
                ws.append([
                    item.get('store_name', ''),
                    item.get('type', ''),
                    item.get('sku', ''),
                    item.get('chinese_name', ''),
                    item.get('quantity', 0),
                    box.get('box_no', ''),
                    box.get('quantity', 0),
                    box.get('specs', ''),
                    item.get('weight', 0),
                    item.get('volume', 0),
                    item.get('remarks', '')
                ])
                
                # 设置单元格样式
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
    
    # 保存为字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def parse_packing_list(df: pd.DataFrame) -> List[Dict]:
    """解析装箱单数据"""
    required_columns = ['店铺', 'SKU', '数量', '箱号', '装箱数量']
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"缺少必要的列: {col}")
    
    # 按店铺和SKU分组处理数据
    result = []
    for (store, sku), group in df.groupby(['店铺', 'SKU']):
        box_quantities = []
        for _, row in group.iterrows():
            box_quantities.append({
                'box_no': str(row['箱号']),
                'quantity': int(row['装箱数量']),
                'specs': str(row.get('规格', ''))
            })
        
        result.append({
            'store_name': store,
            'sku': sku,
            'type': group.iloc[0].get('类型', '普货'),
            'quantity': int(group['数量'].sum()),
            'box_quantities': box_quantities,
            'remarks': group.iloc[0].get('备注', '')
        })
    
    return result 